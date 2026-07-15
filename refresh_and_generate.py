"""
refresh_and_generate.py
───────────────────────
Lit BDD_dashboard.xlsx, régénère les CSVs intermédiaires,
puis lance generate_dashboard.py pour produire dashboard.html.

Usage : double-clic sur update.bat (ou : python refresh_and_generate.py)

Workflow :
  BDD_dashboard.xlsx  →  refresh_and_generate.py  →  5 CSVs  →  generate_dashboard.py  →  dashboard.html
"""
import csv
import os
import re
import subprocess
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))   # test/_internals/
ROOT_DIR = os.path.dirname(SCRIPT_DIR)                    # test/
BASE = SCRIPT_DIR  # Les CSVs et le HTML temp sont generes ici
SRC_XLSX = os.path.join(ROOT_DIR, "BDD_dashboard.xlsx")

# ─── Mappings ─────────────────────────────────────────────────────────────────

# Site → (code contrat, région)
SITE_INFO = {
    "Le Havre":            ("H66NTES281 - 0369-TCS QDR-TRI",  "NNO"),
    "Nantes":              ("H66CAET281 - 9170-CS COUER-TRI", "COU"),
    "Amiens":              ("H66HTVP281 - U397-TRI CS AM-TRI","NNO"),
    "Bègles":              ("H66SVBM281 - U058-BEGLES-TRI",   "SO"),
    "Millau":              ("H66SSMT284 - U446-ECOTRI-TRI",   "SO"),
    "Paris 15":            ("H66IP15281 - 9809-TRI PARI-TRI", "IDF"),
    "Saran":               ("H66CORT282 - U487-TRI IF43-TRI,","COU"),
    "Sevran":              ("H66ISVT281 - U701-SEVRAN-TRI",   "IDF"),
    "Portes les Valences": ("H66RZOS282 - 9171-PORTES VA-TRI","BARA"),
    "Chézy":               ("H66RSAL281 - U310-ALLIER-TRI",   "BARA"),
    "Montpellier":         ("H66SSMT283 - U446-DEMETER-TRI",  "SO"),
}
# Ordre des sites dans les CSVs (match l'ordre original de extract_q1.py)
SITES_ORDER = [
    "Nantes", "Saran", "Amiens", "Paris 15", "Le Havre", "Sevran",
    "Chézy", "Portes les Valences", "Montpellier", "Millau", "Bègles",
]

# Mapping code VE → colonne data_synthese.csv
VE_TO_COL_SYNTHESE = {
    "VE000001": "CA",
    "VE000002": "Prestations_service",
    "VE000015": "Ventes_matieres",
    "VE000024": "Revenus_travaux",
    "VE000035": "Produits_tiers",
    "VE000036": "Couts_ventes_cash",
    "VE000044": "Sous_traitance_externe",
    "VE000045": "Prestations_internes",
    "VE000046": "Prestations_internes_Produit",
    "VE000048": "Sous_traitance_construction",
    "VE000051": "Couts_personnel",
    "VE000067": "Maintenance_courante",
    "VE000071": "Maintenance_obligatoire",
    "VE000077": "Energie",
    "VE000161": "Reactifs_eau",
    "VE000089": "Traitement_sous_produits",
    "VE000097": "Autres_couts_exploitation",
    "VE000108": "Couts_commerciaux",
    "VE000112": "Couts_GA",
    "VE000119": "CDV_personnel_non_cash",
    "VE000121": "Amortissements",
    "VE000187": "Amortissement_IFRIC12",
    "VE000116": "Management_Fees",
}

LIBELLE_TO_COL_SYNTHESE = {
    "Tonnes entrantes":   "Tonnes_entrantes",
    "PNE":                "PNE",
    "Marge Brute Cash":   "Marge_Brute_Cash",
    "EBITDA":             "EBITDA",
    "EBIT Courant":       "EBIT_Courant",
}

VE_TO_DETAIL = {
    "VE000052": "Pers_interne",
    "VE000053": "Pers_externe",
    "VE000054": "Pers_maint_interne",
    "VE000055": "Pers_maint_externe",
    "VE000057": "Pers_avantages",
    "VE000098": "Autr_labo",
    "VE000099": "Autr_equip",
    "VE000100": "Autr_batiment",
    "VE000102": "Autr_taxes",       # VE000102 = Taxes et redevances
    "VE000103": "Autr_assurances",  # VE000103 = Assurances
    "VE000104": "Autr_autres",
}

# Métriques exportées dans data_q1_2026.csv (ordre du dashboard)
# (clé dans BDD_dashboard.xlsx, nom dans le CSV)
# Note : "Marge_brute" est calculée séparément (= CA + VE000036 + VE000050)
Q1_METRICS_EXPORT = [
    ("VE000001",     "CA"),
    ("PNE",          "PNE"),
    ("_MARGE_BRUTE", "Marge_brute"),  # calculée
    ("EBITDA",       "EBITDA"),
    ("EBIT Courant", "EBIT"),
    ("VE000051",     "Personnel"),
    ("VE000077",     "Energie"),
    ("VE000067",     "Maint_courante"),
    ("VE000071",     "Maint_oblig"),
    ("VE000089",     "Traitement_sp"),
    ("VE000097",     "Autres_couts"),
]

YEAR_TO_ANNEE_PL = {
    "R2023": "Réel 2023",
    "R2024": "R 2024",
    "R2025": "Réel 2025",
    "B2026": "B 2026",
}
YEAR_TO_ANNEE_NUM = {
    "R2023": "2023",
    "R2024": "2024",
    "R2025": "2025",
    "B2026": "2026",
}


# ─── Lecture de PL_Annuel : retourne un index {(site, annee, key): valeur} ───
def _to_num(v):
    """Convertit une valeur en nombre si possible, sinon retourne None pour les vides."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def read_pl_annuel(wb, sheet_name):
    """Lit l'onglet PL_Annuel ou Q1_2026 et retourne :
       - site_cols : {site: {annee: col_index}}
       - data : {(site, annee, key): valeur}
       - libelles : {key: libelle_complet}
       - categories : {key: catégorie} (Synthèse / Affiché P&L / Sous-poste / Info / Détail)
    """
    ws = wb[sheet_name]
    site_cols = {}
    last_site = None
    for c in range(4, ws.max_column + 1):
        s = ws.cell(row=1, column=c).value
        if s:
            last_site = s
        if last_site is None:
            continue
        annee = ws.cell(row=2, column=c).value
        site_cols.setdefault(last_site, {})[annee] = c

    data = {}
    libelles = {}
    categories = {}
    for r in range(3, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        libelle = (ws.cell(row=r, column=2).value or "").strip()
        categorie = ws.cell(row=r, column=3).value
        if code and code != "—":
            key = code
            libelle_complet = f"{code} - {libelle}" if libelle else code
        else:
            key = libelle
            libelle_complet = libelle
        libelles[key] = libelle_complet
        categories[key] = categorie
        for site, year_cols in site_cols.items():
            for annee, col_idx in year_cols.items():
                val = ws.cell(row=r, column=col_idx).value
                num_val = _to_num(val)
                if num_val is not None:
                    data[(site, annee, key)] = num_val
    return site_cols, data, libelles, categories


def read_pl_periodique_q1(wb):
    """Lit l'onglet trimestriel (Suivi trimestriel 2026 ou Q1_2026 pour retro-compat).
       Retourne : data = {(site, periode, annee, key): valeur}
       Supporte T1/T2/T3/T4 (BDD) -> Q1/S1/9M/ANN (interne, cumuls).
       Une periode est ignoree si aucune valeur n'y est renseignee.
    """
    # Choix de l'onglet : nouveau nom prioritaire
    if "Suivi trimestriel 2026" in wb.sheetnames:
        ws = wb["Suivi trimestriel 2026"]
    elif "Q1_2026" in wb.sheetnames:
        ws = wb["Q1_2026"]
    else:
        return None

    # Detection : L2 col 4 doit valoir 'T1', 'Q1' (ancienne convention) ou similaire
    l2_c4 = ws.cell(row=2, column=4).value
    if l2_c4 not in ("T1", "Q1"):
        # Ancien format (2 lignes header) : fallback vers read_pl_annuel
        return None

    # Mapping libelle affiche -> code interne (cumuls a fin de periode)
    # T1 = cumul fin T1 (jan-mars) ; T2 = cumul S1 (jan-juin) ;
    # T3 = cumul 9M (jan-sept)     ; T4 = cumul annuel (jan-dec)
    PERIOD_ALIAS = {
        "T1": "Q1", "T2": "S1", "T3": "9M", "T4": "ANN",
        "Q1": "Q1", "S1": "S1",  # retro-compat
    }

    # Parcourir les colonnes pour construire {site: {periode: {annee: col_idx}}}
    site_period_cols = {}
    last_site = None
    last_period = None
    for c in range(4, ws.max_column + 1):
        s = ws.cell(row=1, column=c).value
        if s:
            last_site = s
        p = ws.cell(row=2, column=c).value
        if p:
            # Traduit T1/T2 (BDD) vers Q1/S1 (interne) pour compat JS
            last_period = PERIOD_ALIAS.get(p, p)
        if last_site is None or last_period is None:
            continue
        annee = ws.cell(row=3, column=c).value
        site_period_cols.setdefault(last_site, {}).setdefault(last_period, {})[annee] = c

    data = {}
    for r in range(4, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        libelle = (ws.cell(row=r, column=2).value or "").strip()
        if code and code != "—":
            key = code
        else:
            key = libelle
        if not key:
            continue
        for site, periods in site_period_cols.items():
            for period, year_cols in periods.items():
                for annee, col_idx in year_cols.items():
                    val = ws.cell(row=r, column=col_idx).value
                    num_val = _to_num(val)
                    if num_val is not None:
                        data[(site, period, annee, key)] = num_val

    return data


# ─── 1. data_synthese.csv ─────────────────────────────────────────────────────
def write_data_synthese(wb):
    _, data, _, _ = read_pl_annuel(wb, "PL_Annuel")
    headers = [
        "Site","Code","Annee_PL","Tonnes_entrantes","CA","Prestations_service",
        "Ventes_matieres","Revenus_travaux","Produits_tiers","Couts_ventes_cash",
        "Sous_traitance_externe","Prestations_internes","Prestations_internes_Produit","Sous_traitance_construction",
        "PNE","Couts_personnel","Maintenance_courante","Maintenance_obligatoire",
        "Energie","Reactifs_eau","Traitement_sous_produits","Autres_couts_exploitation",
        "Marge_Brute_Cash","Couts_commerciaux","Couts_GA","EBITDA",
        "CDV_personnel_non_cash","Amortissements","Amortissement_IFRIC12","EBIT_Courant",
        "Annee","Region","Management_Fees"
    ]
    rows = []
    for site in SITES_ORDER:
        code_contrat, region = SITE_INFO[site]
        for annee in ["R2023","R2024","R2025","B2026"]:
            row = {h: "" for h in headers}
            row["Site"] = site
            row["Code"] = code_contrat
            row["Annee_PL"] = YEAR_TO_ANNEE_PL[annee]
            row["Annee"] = YEAR_TO_ANNEE_NUM[annee]
            row["Region"] = region
            # Remplir les colonnes par VE code
            for ve, col in VE_TO_COL_SYNTHESE.items():
                v = data.get((site, annee, ve))
                row[col] = v if v is not None else ""
            # Remplir les colonnes par libellé
            for lib, col in LIBELLE_TO_COL_SYNTHESE.items():
                v = data.get((site, annee, lib))
                if v is None:
                    # Tenter avec préfixe "[R0100-EW-109] Tonnes entrantes"
                    if lib == "Tonnes entrantes":
                        v = data.get((site, annee, "[R0100-EW-109] Tonnes entrantes"))
                row[col] = v if v is not None else ""
            rows.append(row)
    _write_csv(os.path.join(BASE, "data_synthese.csv"), headers, rows)
    print(f"  [OK] data_synthese.csv         ({len(rows)} lignes)")


# ─── 2. data_synthese_eur_t.csv ───────────────────────────────────────────────
def write_data_synthese_eur_t(wb):
    _, data, libelles, categories = read_pl_annuel(wb, "PL_Annuel")
    # Calcul des Tonnes_entrantes pour chaque (site, annee)
    tonnage = {}
    for site in SITES_ORDER:
        for annee in ["R2023","R2024","R2025"]:
            # On cherche par libellé "Tonnes entrantes" (différents formats)
            t = None
            for lib in ["[R0100-EW-109] Tonnes entrantes", "Tonnes entrantes"]:
                v = data.get((site, annee, lib))
                if v is not None:
                    t = v
                    break
            tonnage[(site, annee)] = t

    headers = ["Site","Code","Annee","Metrique","Valeur_EUR_t"]
    rows = []
    # Trouver toutes les clés présentes (lignes du P&L)
    all_keys = sorted({k for (_, _, k) in data.keys()})
    # On veut un ordre logique : ordre Nantes
    nantes_keys = [k for (s, _, k) in data.keys() if s == "Nantes"]
    seen = set()
    ordered_keys = []
    for k in nantes_keys:
        if k not in seen:
            ordered_keys.append(k)
            seen.add(k)
    for k in all_keys:
        if k not in seen:
            ordered_keys.append(k)
            seen.add(k)

    for site in SITES_ORDER:
        code_contrat, _ = SITE_INFO[site]
        for annee in ["R2023","R2024","R2025"]:
            tn = tonnage.get((site, annee))
            for key in ordered_keys:
                # Skip codes analytiques (catégorie "Détail")
                cat = categories.get(key)
                if cat == "Détail":
                    continue
                v = data.get((site, annee, key))
                if v is None:
                    continue
                libelle = libelles.get(key, key)
                # Pour "Tonnes entrantes", on écrit la valeur brute
                if "Tonnes" in libelle:
                    val_eur_t = v
                else:
                    val_eur_t = v / tn if tn else ""
                rows.append({
                    "Site": site,
                    "Code": code_contrat,
                    "Annee": annee,
                    "Metrique": libelle,
                    "Valeur_EUR_t": val_eur_t if val_eur_t != "" else "",
                })
    _write_csv(os.path.join(BASE, "data_synthese_eur_t.csv"), headers, rows)
    print(f"  [OK] data_synthese_eur_t.csv   ({len(rows)} lignes)")


# ─── 3. data_detail_pl.csv ────────────────────────────────────────────────────
def write_data_detail_pl(wb):
    _, data, _, _ = read_pl_annuel(wb, "PL_Annuel")
    headers = ["Site","Annee"] + list(VE_TO_DETAIL.values())
    rows = []
    for site in SITES_ORDER:
        for annee in ["R2023","R2024","R2025"]:
            row = {"Site": site, "Annee": annee}
            for ve, col in VE_TO_DETAIL.items():
                v = data.get((site, annee, ve))
                row[col] = v if v is not None else ""
            rows.append(row)
    _write_csv(os.path.join(BASE, "data_detail_pl.csv"), headers, rows)
    print(f"  [OK] data_detail_pl.csv        ({len(rows)} lignes)")


# ─── 4. data_q1_2026.csv ──────────────────────────────────────────────────────
# Definit toutes les periodes possibles + comment les calculer.
# Cumuls (T1/T2/T3/T4) = valeurs brutes de la BDD (codes Q1/S1/9M/ANN).
# Trimestres isoles (Q2/Q3/Q4) = calcul par difference.
PERIODS_DEF = [
    # (code_interne, cumul_bdd_source, cumul_precedent_ou_none, isolé?)
    ("Q1",  "Q1",  None,  False),   # T1 cumul jan-mars (= T1 isole aussi)
    ("Q2",  "S1",  "Q1",  True),    # T2 isole = S1 - Q1 (avr-juin)
    ("S1",  "S1",  None,  False),   # Cumul S1 (jan-juin)
    ("Q3",  "9M",  "S1",  True),    # T3 isole = 9M - S1 (juil-sept)
    ("9M",  "9M",  None,  False),   # Cumul 9M (jan-sept)
    ("Q4",  "ANN", "9M",  True),    # T4 isole = ANN - 9M (oct-dec)
    ("ANN", "ANN", None,  False),   # Cumul annuel
]


def write_data_q1(wb):
    """Genere data_q1_2026.csv avec toutes les periodes disponibles.
       Structure : Site, Metrique, Periode, R2025, R2026, B2026.
       Periodes possibles : Q1, Q2, S1, Q3, 9M, Q4, ANN.
       Une periode n'est ecrite que si la source BDD correspondante est renseignee.
    """
    periodic_data = read_pl_periodique_q1(wb)

    # Detection des cumuls presents (au moins une valeur non vide dans le cumul)
    cumuls_present = set()
    for (site, period, annee, key), v in periodic_data.items():
        if v is not None and v != "":
            cumuls_present.add(period)

    # On genere une periode si son cumul source est present dans la BDD
    active_periods = [p for p in PERIODS_DEF if p[1] in cumuls_present]
    active_codes = [p[0] for p in active_periods]

    def get_val(site, period_code, annee, key):
        """Recupere la valeur cumul brut ou calcul isole."""
        pdef = next((p for p in PERIODS_DEF if p[0] == period_code), None)
        if pdef is None:
            return None
        _, cumul_source, cumul_prev, is_isole = pdef
        v_cur = periodic_data.get((site, cumul_source, annee, key))
        if not is_isole:
            return v_cur
        # Isole : difference avec le cumul precedent
        v_prev = periodic_data.get((site, cumul_prev, annee, key))
        if v_cur is None or v_prev is None:
            return None
        return v_cur - v_prev

    headers = ["Site", "Metrique", "Periode", "R2025", "R2026", "B2026"]
    rows = []

    for site in SITES_ORDER:
        for period_code in active_codes:
            for key, metric_name in Q1_METRICS_EXPORT:
                row = {"Site": site, "Metrique": metric_name, "Periode": period_code}
                for annee in ["R2025", "R2026", "B2026"]:
                    if key == "_MARGE_BRUTE":
                        ca  = get_val(site, period_code, annee, "VE000001")
                        v36 = get_val(site, period_code, annee, "VE000036")
                        v50 = get_val(site, period_code, annee, "VE000050")
                        if None in (ca, v36, v50):
                            v = ""
                        else:
                            v = ca + v36 + v50
                    else:
                        v = get_val(site, period_code, annee, key)
                        if v is None:
                            v = ""
                    row[annee] = v
                rows.append(row)
    _write_csv(os.path.join(BASE, "data_q1_2026.csv"), headers, rows)
    print(f"  [OK] data_q1_2026.csv          ({len(rows)} lignes, periodes actives: {'/'.join(active_codes)})")


# ─── 5. data_kpi_techniques.csv ───────────────────────────────────────────────
def write_data_kpi(wb):
    ws = wb["KPI_Techniques"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            v = ws.cell(row=r, column=ci).value
            row[h] = v if v is not None else ""
        if row.get("Site"):
            rows.append(row)
    _write_csv(os.path.join(BASE, "data_kpi_techniques.csv"), headers, rows)
    print(f"  [OK] data_kpi_techniques.csv   ({len(rows)} lignes)")


def _write_csv(path, headers, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for row in rows:
            w.writerow(row)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(SRC_XLSX):
        print(f"[ERREUR] {SRC_XLSX} introuvable.")
        sys.exit(1)

    print(f"\n[1/2] Lecture de BDD_dashboard.xlsx et generation des CSVs...\n")
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    write_data_synthese(wb)
    write_data_synthese_eur_t(wb)
    write_data_detail_pl(wb)
    write_data_q1(wb)
    write_data_kpi(wb)

    print(f"\n[2/2] Generation du dashboard...\n")
    result = subprocess.run(
        [sys.executable, os.path.join(SCRIPT_DIR, "generate_dashboard.py")],
        cwd=SCRIPT_DIR,
    )
    if result.returncode != 0:
        print("\n[ERREUR] La generation du dashboard a echoue.")
        sys.exit(result.returncode)

    # Copier le dashboard.html vers le dossier racine (visible pour l'utilisateur)
    import shutil
    src_html = os.path.join(SCRIPT_DIR, "dashboard.html")
    dst_html = os.path.join(ROOT_DIR, "dashboard.html")
    if os.path.exists(src_html):
        shutil.copy(src_html, dst_html)

    print(f"\n[OK] Tout est a jour ! Ouvrez dashboard.html dans votre navigateur.")


if __name__ == "__main__":
    main()
