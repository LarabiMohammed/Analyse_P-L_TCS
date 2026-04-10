"""
extract_q1.py — Extraction des données Q1 2026 depuis P&L Tri Réel 03-2026.xlsx

Produit data_q1_2026.csv avec, pour chacun des 11 sites du dashboard :
  CA, PNE, Marge_brute, EBITDA, EBIT, Personnel, Energie,
  Maint_courante, Maint_oblig, Traitement_sp, Autres_couts

Colonnes : Site, Metrique, R2025, R2026, B2026 (valeurs en euros bruts)

Marge_brute = CA + VE000036 + VE000050  (calculée, pas une ligne directe)
Plusieurs onglets par site → agrégation par somme.
"""

import csv
import openpyxl

# ── 1. Mapping explicite site_dashboard → liste de codes onglets Excel ─────────
# Les codes correspondent au début des noms d'onglets dans le fichier P&L.
SITE_SHEETS = {
    "Nantes":            ["H66CAET281"],
    "Saran":             ["H66CORT282"],
    "Amiens":            ["H66HTVP281"],
    "Paris 15":          ["H66IP15281"],
    "Le Havre":          ["H66NTES281"],
    "Sevran":            ["H66ISET282", "H66ISVT281", "H66ISVT283"],  # 3 onglets
    "Chézy":             ["H66RSAL281"],
    "Portes les Valences": ["H66RZOS282"],
    "Montpellier":       ["H66SSMT261", "H66SSMT281", "H66SSMT283"],  # 3 onglets
    "Millau":            ["H66SSMT284"],
    "Bègles":            ["H66SVBM281"],
}

# ── 2. Lignes à chercher dans colonne A de chaque onglet ────────────────────────
# Clé interne → (type de match, label court)
# type "startswith" : col_a.strip().startswith(key)
# type "exact"      : col_a.strip() == key
METRICS_KEYS = [
    ("VE000001", "startswith", "CA"),
    ("VE000036", "startswith", "_VE000036"),     # intermédiaire pour Marge_brute
    ("VE000050", "startswith", "_VE000050"),     # intermédiaire pour Marge_brute
    ("PNE",      "exact",      "PNE"),
    ("EBITDA",   "exact",      "EBITDA"),
    ("EBIT Courant", "exact",  "EBIT"),
    ("VE000051", "startswith", "Personnel"),
    ("VE000077", "startswith", "Energie"),
    ("VE000067", "startswith", "Maint_courante"),
    ("VE000071", "startswith", "Maint_oblig"),
    ("VE000089", "startswith", "Traitement_sp"),
    ("VE000097", "startswith", "Autres_couts"),
]

def safe_float(v):
    """Convertit une valeur en float, None si manquante ou non numérique."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def extract_sheet(ws):
    """
    Parcourt un onglet Excel et renvoie un dict {label_court: [R2025, R2026, B2026]}.
    Colonne A = libellé, B = R2025, C = R2026, D = Budget 2026.
    """
    results = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        col_a = row[0]
        if not isinstance(col_a, str):
            continue
        stripped = col_a.strip()
        for key, match_type, label in METRICS_KEYS:
            if match_type == "startswith" and stripped.startswith(key):
                # Ne prendre que la ligne de subtotal (pas les sous-lignes)
                if label not in results:
                    r25 = safe_float(row[1]) if len(row) > 1 else None
                    r26 = safe_float(row[2]) if len(row) > 2 else None
                    b26 = safe_float(row[3]) if len(row) > 3 else None
                    results[label] = [r25, r26, b26]
                break
            elif match_type == "exact" and stripped == key:
                if label not in results:
                    r25 = safe_float(row[1]) if len(row) > 1 else None
                    r26 = safe_float(row[2]) if len(row) > 2 else None
                    b26 = safe_float(row[3]) if len(row) > 3 else None
                    results[label] = [r25, r26, b26]
                break
    return results


def add_vals(a, b):
    """Somme deux listes [R2025, R2026, B2026] en gérant les None."""
    result = []
    for x, y in zip(a, b):
        if x is None and y is None:
            result.append(None)
        else:
            result.append((x or 0.0) + (y or 0.0))
    return result


def aggregate_site(wb, sheet_prefixes):
    """
    Agrège les données de plusieurs onglets pour un même site (somme des valeurs).
    sheet_prefixes : liste de codes (ex. ["H66ISET282", "H66ISVT281", "H66ISVT283"])
    Renvoie un dict {label_court: [R2025, R2026, B2026]}
    """
    totals = {}
    matched_sheets = 0
    for sheet_name in wb.sheetnames:
        for prefix in sheet_prefixes:
            if sheet_name.startswith(prefix):
                ws = wb[sheet_name]
                sheet_data = extract_sheet(ws)
                matched_sheets += 1
                for label, vals in sheet_data.items():
                    if label in totals:
                        totals[label] = add_vals(totals[label], vals)
                    else:
                        totals[label] = vals[:]
                break  # évite de compter deux fois si deux préfixes matchent

    if matched_sheets == 0:
        print(f"  ⚠ Aucun onglet trouvé pour les préfixes {sheet_prefixes}")
    return totals


def compute_marge_brute(data):
    """
    Marge_brute = CA + VE000036 + VE000050
    Les trois sont positifs pour CA, négatifs pour VE000036 et VE000050.
    """
    result = [None, None, None]
    if "CA" in data and "_VE000036" in data and "_VE000050" in data:
        for i in range(3):
            v_ca  = (data["CA"][i] or 0.0)
            v_036 = (data["_VE000036"][i] or 0.0)
            v_050 = (data["_VE000050"][i] or 0.0)
            if data["CA"][i] is not None:
                result[i] = v_ca + v_036 + v_050
    return result


def main():
    wb = openpyxl.load_workbook(
        "P&L Tri Réel 03-2026.xlsx", read_only=True, data_only=True
    )

    rows_out = []

    for site_name, prefixes in SITE_SHEETS.items():
        print(f"Traitement : {site_name}  (onglets : {prefixes})")
        data = aggregate_site(wb, prefixes)

        # Marge brute calculée
        data["Marge_brute"] = compute_marge_brute(data)

        # Métriques à exporter (dans l'ordre souhaité dans le dashboard)
        metrics_export = [
            ("CA",            "CA"),
            ("PNE",           "PNE"),
            ("Marge_brute",   "Marge_brute"),
            ("EBITDA",        "EBITDA"),
            ("EBIT",          "EBIT"),
            ("Personnel",     "Personnel"),
            ("Energie",       "Energie"),
            ("Maint_courante","Maint_courante"),
            ("Maint_oblig",   "Maint_oblig"),
            ("Traitement_sp", "Traitement_sp"),
            ("Autres_couts",  "Autres_couts"),
        ]

        for label_in_data, label_csv in metrics_export:
            vals = data.get(label_in_data, [None, None, None])
            rows_out.append({
                "Site":     site_name,
                "Metrique": label_csv,
                "R2025":    round(vals[0], 2) if vals[0] is not None else "",
                "R2026":    round(vals[1], 2) if vals[1] is not None else "",
                "B2026":    round(vals[2], 2) if vals[2] is not None else "",
            })

    # Écriture CSV
    output_file = "data_q1_2026.csv"
    with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["Site", "Metrique", "R2025", "R2026", "B2026"])
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"\nOK {len(rows_out)} lignes ecrites dans {output_file}")
    print(f"  {len(SITE_SHEETS)} sites x 11 metriques = {len(SITE_SHEETS)*11} attendu")


if __name__ == "__main__":
    main()
