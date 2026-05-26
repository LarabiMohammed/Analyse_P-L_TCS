"""
build_master_excel.py
─────────────────────
Génère BDD_dashboard.xlsx — fichier Excel maître pour le dashboard.

Lit DIRECTEMENT le fichier source 'Copie de P&L TRI R2023-R2024-R2025-B2026.xlsx'
et reproduit TOUTES les lignes du P&L (233 par site), en gardant la hiérarchie
et l'indentation d'origine.

Structure générée :
  - _README          : description de la structure
  - PL_Annuel        : toutes les lignes du P&L, color-codées
  - Q1_2026          : comptes Q1 2026 (depuis CSV existant)
  - KPI_Techniques   : KPIs opérationnels (depuis CSV existant)

Code couleurs PL_Annuel :
  - 🟦 Bleu marine  : soldes intermédiaires (CA, PNE, Marge, EBITDA, EBIT)
  - 🔷 Bleu clair   : lignes affichées dans le dashboard
  - 🟢 Vert clair   : sous-postes (drill-down Personnel / Autres coûts)
  - 🟡 Jaune clair  : composantes / lignes pour info
  - ⬜ Gris très clair : sous-codes analytiques (66000xxx, 665Xxxxx)

Usage : python build_master_excel.py
"""
import csv
import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(BASE, "BDD_dashboard.xlsx")
SRC_FILE = os.path.join(BASE, "Copie de P&L TRI R2023-R2024-R2025-B2026.xlsx")
SRC_Q1_FILE = os.path.join(BASE, "P&L Tri Réel 03-2026.xlsx")

# Mapping onglet Excel source → nom de site
SITE_SHEETS = [
    # (sheet_name, site_name)
    ("H66NTES281 - 0369-TCS QDR-TRI",     "Le Havre"),
    ("H66CAET281 - 9170-CS COUER-TRI",    "Nantes"),
    ("H66HTVP281 - U397-TRI CS AM-TRI",   "Amiens"),
    ("H66SVBM281 - U058-BEGLES-TRI",      "Bègles"),
    ("H66SSMT284 - U446-ECOTRI-TRI",      "Millau"),
    ("H66IP15281 - 9809-TRI PARI-TRI",    "Paris 15"),
    ("H66CORT282 - U487-TRI IF43-TRI,",   "Saran"),
    ("H66ISVT281 - U701-SEVRAN-TRI",      "Sevran"),
    ("H66RZOS282 - 9171-PORTES VA-TRI",   "Portes les Valences"),
    ("H66RSAL281 - U310-ALLIER-TRI",      "Chézy"),
    ("H66SSMT283 - U446-DEMETER-TRI",     "Montpellier"),
]

# Colonnes de l'Excel source annuel à conserver : (label, index_colonne 1-based)
# B=2 Réel 2023, C=3 R 2024, D=4 B 2025 (skip), E=5 Réel 2025, F=6 B 2026
YEAR_COLS = [("R2023", 2), ("R2024", 3), ("R2025", 5), ("B2026", 6)]

# Corrections manuelles à appliquer après lecture du fichier source
# Format : {(site, code/libellé, année): valeur_corrigée}
# Note : ces corrections existent parce que le fichier source contient des erreurs
#        que l'utilisateur a corrigées à la main dans les CSVs d'origine.
CORRECTIONS_PL_ANNUEL = {
    ("Bègles", "EBITDA", "R2023"): -1875548.81,
}

# Colonnes de l'Excel source Q1 à conserver : (label, index_colonne 1-based)
# B=2 R 2025, C=3 R 2026, D=4 Budget 2026
Q1_YEAR_COLS = [("R2025", 2), ("R2026", 3), ("B2026", 4)]

# Catégories pour le code couleur
# Match par préfixe / mot-clé dans le libellé (sans les espaces d'indentation)
SYNTH_KEYS = [
    "VE000001 - Chiffre d'affaires",
    "PNE",
    "Marge Brute Cash",
    "EBITDA",
    "EBIT Courant",
]

# Lignes affichées dans le P&L principal du dashboard
PRINCIPAL_KEYS = [
    "VE000051 - Coûts de personnel",
    "VE000067 - Entretien & maintenance courante",
    "VE000071 - Maint. Obligatoire",
    "VE000077 - Coût des énergies",
    "VE000089 - Traitement et évacuation",
    "VE000097 - Autres coûts d'exploitation",
    "VE000108 - Coûts Commerciaux",
    "VE000112 - Coûts Généraux",
    "VE000121 - CDV - Autres amortissements",
    "VE000187 - Amortissement Actif Incorporel",
]

# Sous-postes du drill-down (Personnel et Autres coûts)
SOUS_POSTE_KEYS = [
    "VE000052 - Personnel exploitation - interne",
    "VE000053 - Personnel exploitation - externe",
    "VE000054 - Personnel maintenance - interne",
    "VE000055 - Personnel maintenance - externe",
    "VE000057 - Autres avantages",
    "VE000098 - Laboratoire",
    "VE000099 - Équipements mobiles",
    "VE000099 - Equipements mobiles",
    "VE000100 - Bâtiments",
    "VE000100 - Batiments",
    "VE000102 - Taxes",        # VE000102 = Taxes et redevances
    "VE000103 - Assurances",   # VE000103 = Assurances
    "VE000104 - Autres",
]

# Styles par catégorie
CAT_STYLE = {
    "SYNTH":      {"fill": "003A63", "font": "FFFFFF", "bold": True,  "label": "Synthèse"},
    "PRINCIPAL":  {"fill": "BFDBFE", "font": "1E3A8A", "bold": True,  "label": "Affiché P&L"},
    "SOUS_POSTE": {"fill": "BBF7D0", "font": "166534", "bold": False, "label": "Sous-poste"},
    "INFO":       {"fill": "FEF3C7", "font": "92400E", "bold": False, "label": "Info"},
    "ANALYTIQUE": {"fill": "F3F4F6", "font": "6B7280", "bold": False, "label": "Détail"},
}

# Styles communs
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="003A63", end_color="003A63", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
SITE_HEADER_FILL = PatternFill(start_color="0E5A8A", end_color="0E5A8A", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="003A63")
SECTION_FONT = Font(bold=True, size=12, color="003A63")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def read_csv(path):
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def try_num(v):
    if v == "" or v is None:
        return None
    try:
        if isinstance(v, (int, float)):
            return v
        s = str(v)
        if "." not in s and "e" not in s.lower():
            return int(s)
        return float(s)
    except (ValueError, TypeError):
        return v


def get_indent(libelle_raw):
    """Compte les espaces de tête pour déterminer le niveau d'indentation."""
    if libelle_raw is None:
        return 0
    s = str(libelle_raw)
    n = len(s) - len(s.lstrip(" "))
    return n // 2  # par convention : 2 espaces = 1 niveau


def classify_line(libelle_clean, indent):
    """Détermine la catégorie d'une ligne du P&L."""
    if not libelle_clean:
        return "INFO"

    # 1. Synthèses
    for key in SYNTH_KEYS:
        if libelle_clean.startswith(key):
            return "SYNTH"

    # 2. Sous-postes (drill-down)
    for key in SOUS_POSTE_KEYS:
        if libelle_clean.startswith(key):
            return "SOUS_POSTE"

    # 3. Lignes principales du dashboard
    for key in PRINCIPAL_KEYS:
        if libelle_clean.startswith(key):
            return "PRINCIPAL"

    # 4. Détail analytique (codes 66xxx, 664xxx, 665Xxxxx, etc.)
    if re.match(r"^(66\d{6,}|664[A-Z]+\d+|665[A-Z]\d+)\b", libelle_clean):
        return "ANALYTIQUE"

    # 5. Tonnage et autres mesures d'opération
    if libelle_clean.startswith("[R") or "Tonnes" in libelle_clean or "Tonnage" in libelle_clean:
        return "INFO"

    # 6. Reste : info
    return "INFO"


# ─── Construction de PL_Annuel à partir du fichier source ─────────────────────
def line_key(libelle_clean):
    """Extrait une clé de matching à partir d'un libellé (code VE/66xxx ou libellé)."""
    # Match préfixe code analytique
    m = re.match(r"^([A-Z0-9]+(?:-[A-Z0-9]+)?)\s*-", libelle_clean)
    if m:
        return m.group(1)  # ex. "VE000051", "66000096", "665X0543"
    # Sinon : libellé entier (pour PNE, EBITDA, Marge Brute Cash, EBIT Courant...)
    return libelle_clean


def write_pl_annuel(wb):
    """Lit le fichier source et reproduit toutes les lignes du P&L.

    Chaque site ayant sa propre structure, on matche les lignes par CODE/LIBELLÉ
    et pas par numéro de ligne.
    """
    print("  Lecture du fichier source...")
    src = openpyxl.load_workbook(SRC_FILE, read_only=True, data_only=True)

    # 1. Lire toutes les lignes de chaque site, indexées par leur clé (code/libellé)
    # site_lines[site] = {key: (libelle_raw, libelle_clean, indent, year_values)}
    # site_order[site] = [keys dans l'ordre où ils apparaissent]
    site_lines = {}
    site_order = {}
    for sheet_name, site_name in SITE_SHEETS:
        if sheet_name not in src.sheetnames:
            print(f"  [WARN] Onglet '{sheet_name}' introuvable pour {site_name}")
            continue
        ws = src[sheet_name]
        site_lines[site_name] = {}
        site_order[site_name] = []
        for row in ws.iter_rows(min_col=1, max_col=6, values_only=True):
            v = row[0]
            if v is None or not str(v).strip():
                continue
            libelle_raw = str(v)
            libelle_clean = libelle_raw.strip()
            indent = get_indent(libelle_raw)
            key = line_key(libelle_clean)
            yr_vals = {}
            for annee, col_idx in YEAR_COLS:
                if col_idx - 1 < len(row):
                    yr_vals[annee] = row[col_idx - 1]
            # En cas de doublon de clé (rare), garde la première occurrence
            if key not in site_lines[site_name]:
                site_lines[site_name][key] = (libelle_raw, libelle_clean, indent, yr_vals)
                site_order[site_name].append(key)
        print(f"  [OK] {site_name:<22} ({len(site_lines[site_name])} lignes)")

    src.close()

    # Appliquer les corrections manuelles
    for (site, key_or_libelle, annee), corrected_val in CORRECTIONS_PL_ANNUEL.items():
        if site in site_lines and key_or_libelle in site_lines[site]:
            libelle_raw, libelle_clean, indent, yr_vals = site_lines[site][key_or_libelle]
            yr_vals[annee] = corrected_val
            site_lines[site][key_or_libelle] = (libelle_raw, libelle_clean, indent, yr_vals)
            print(f"  [CORR] {site} {key_or_libelle} {annee} -> {corrected_val}")

    # 2. Construire l'ordre unifié : union des clés à partir de Nantes (réf)
    # puis ajouter les lignes spécifiques aux autres sites
    REF_SITE = "Nantes" if "Nantes" in site_order else list(site_order.keys())[0]
    unified_order = list(site_order[REF_SITE])
    unified_set = set(unified_order)

    # Ajouter les lignes uniques des autres sites (à la fin pour préserver l'ordre)
    for site, keys in site_order.items():
        if site == REF_SITE:
            continue
        for k in keys:
            if k not in unified_set:
                unified_order.append(k)
                unified_set.add(k)

    print(f"  Structure unifiée : {len(unified_order)} lignes (réf : {REF_SITE})")

    # 3. Pour chaque ligne unifiée, récupérer le libellé "canonique"
    # (on prend le libellé du site de référence s'il existe, sinon du premier site qui l'a)
    # structure = [(libelle_raw, libelle_clean, indent, key)]
    structure = []
    for key in unified_order:
        # Trouver une définition du libellé
        ref_def = site_lines.get(REF_SITE, {}).get(key)
        if ref_def is None:
            # Chercher dans les autres sites
            for site in site_order:
                if key in site_lines[site]:
                    ref_def = site_lines[site][key]
                    break
        if ref_def:
            libelle_raw, libelle_clean, indent, _ = ref_def
            structure.append((libelle_raw, libelle_clean, indent, key))

    # 4. Créer l'onglet PL_Annuel
    ws = wb.create_sheet(title="PL_Annuel")

    # En-tête : Code | Libellé | Catégorie | (Site1: R2023, R2024, R2025, B2026) | ...
    ws.cell(row=1, column=1, value="Code").alignment = HEADER_ALIGN
    ws.cell(row=1, column=2, value="Libellé P&L").alignment = HEADER_ALIGN
    ws.cell(row=1, column=3, value="Catégorie").alignment = HEADER_ALIGN
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    site_names_in_order = [s for (_, s) in SITE_SHEETS if s in site_lines]
    col = 4
    for site in site_names_in_order:
        ws.cell(row=1, column=col, value=site)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(YEAR_COLS) - 1)
        col += len(YEAR_COLS)

    # Ligne 2 : années
    col = 4
    for _ in site_names_in_order:
        for annee, _ in YEAR_COLS:
            ws.cell(row=2, column=col, value=annee)
            col += 1

    # Mise en forme en-têtes
    total_cols = 3 + len(site_names_in_order) * len(YEAR_COLS)
    for col_idx in range(1, total_cols + 1):
        for row_idx_h in (1, 2):
            cell = ws.cell(row=row_idx_h, column=col_idx)
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            if row_idx_h == 1 and col_idx >= 4:
                cell.fill = SITE_HEADER_FILL
            else:
                cell.fill = HEADER_FILL

    # 5. Écrire les données : une ligne par entrée de structure
    for i, (libelle_raw, libelle_clean, indent, key) in enumerate(structure):
        excel_row = i + 3
        cat = classify_line(libelle_clean, indent)
        style = CAT_STYLE[cat]
        fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        font = Font(color=style["font"], bold=style["bold"], size=10)

        # Extraire le code analytique
        code_match = re.match(r"^([A-Z0-9]+)\s*-", libelle_clean)
        code = code_match.group(1) if code_match else ""

        # Libellé sans le code
        if code and libelle_clean.startswith(code):
            libelle_sans_code = libelle_clean[len(code):].lstrip(" -").strip()
        else:
            libelle_sans_code = libelle_clean

        indent_prefix = "  " * indent

        c = ws.cell(row=excel_row, column=1, value=code if code else "—")
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=excel_row, column=2, value=indent_prefix + libelle_sans_code)
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=excel_row, column=3, value=style["label"])
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        # Valeurs : on cherche la ligne par sa clé dans le dict de chaque site
        col_v = 4
        for site in site_names_in_order:
            line_def = site_lines.get(site, {}).get(key)
            yr_vals = line_def[3] if line_def else {}
            for annee, _ in YEAR_COLS:
                val = try_num(yr_vals.get(annee))
                cell = ws.cell(row=excel_row, column=col_v, value=val)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = THIN_BORDER
                cell.number_format = "#,##0;-#,##0;—"
                col_v += 1

    # 5. Largeurs et figement
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    for col_idx in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "D3"

    # Activer le filtre auto sur les colonnes A, B, C
    ws.auto_filter.ref = f"A2:C{len(structure) + 2}"

    print(f"  [OK] PL_Annuel : {len(structure)} lignes × {len(site_names_in_order)} sites × {len(YEAR_COLS)} années")


# ─── Sheet Q1_2026 (depuis le fichier source) ─────────────────────────────────
def write_q1(wb):
    """Lit P&L Tri Réel 03-2026.xlsx et reproduit toutes les lignes du P&L Q1."""
    if not os.path.exists(SRC_Q1_FILE):
        print(f"  [WARN] Fichier Q1 source introuvable : {SRC_Q1_FILE}")
        return

    print("  Lecture du fichier Q1 source...")
    src = openpyxl.load_workbook(SRC_Q1_FILE, read_only=True, data_only=True)

    # 1. Lire toutes les lignes de chaque site (matching par clé code/libellé)
    site_lines = {}
    site_order = {}
    for sheet_name, site_name in SITE_SHEETS:
        actual_sheet = None
        if sheet_name in src.sheetnames:
            actual_sheet = sheet_name
        else:
            base_name = sheet_name.rstrip(",")
            for s in src.sheetnames:
                if s.startswith(base_name) or s.startswith(sheet_name.split(" - ")[0]):
                    actual_sheet = s
                    break
        if actual_sheet is None:
            print(f"  [WARN] Onglet Q1 introuvable pour {site_name}")
            continue

        ws = src[actual_sheet]
        site_lines[site_name] = {}
        site_order[site_name] = []
        for row in ws.iter_rows(min_col=1, max_col=7, values_only=True):
            v = row[0]
            if v is None or not str(v).strip():
                continue
            libelle_raw = str(v)
            libelle_clean = libelle_raw.strip()
            indent = get_indent(libelle_raw)
            key = line_key(libelle_clean)
            yr_vals = {}
            for annee, col_idx in Q1_YEAR_COLS:
                if col_idx - 1 < len(row):
                    yr_vals[annee] = row[col_idx - 1]
            if key not in site_lines[site_name]:
                site_lines[site_name][key] = (libelle_raw, libelle_clean, indent, yr_vals)
                site_order[site_name].append(key)
        print(f"  [OK] {site_name:<22} ({len(site_lines[site_name])} lignes)")

    src.close()

    # 2. Ordre unifié : Nantes comme référence, puis ajout des lignes uniques
    REF_SITE = "Nantes" if "Nantes" in site_order else list(site_order.keys())[0]
    unified_order = list(site_order[REF_SITE])
    unified_set = set(unified_order)
    for site, keys in site_order.items():
        if site == REF_SITE:
            continue
        for k in keys:
            if k not in unified_set:
                unified_order.append(k)
                unified_set.add(k)

    # 3. Structure canonique
    structure = []
    for key in unified_order:
        ref_def = site_lines.get(REF_SITE, {}).get(key)
        if ref_def is None:
            for site in site_order:
                if key in site_lines[site]:
                    ref_def = site_lines[site][key]
                    break
        if ref_def:
            libelle_raw, libelle_clean, indent, _ = ref_def
            structure.append((libelle_raw, libelle_clean, indent, key))

    # 4. Créer l'onglet Q1_2026 avec la même mise en forme que PL_Annuel
    ws = wb.create_sheet(title="Q1_2026")

    site_names_in_order = [s for (_, s) in SITE_SHEETS if s in site_lines]

    # Ligne 1 : sites fusionnés
    ws.cell(row=1, column=1, value="Code")
    ws.cell(row=1, column=2, value="Libellé P&L")
    ws.cell(row=1, column=3, value="Catégorie")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    col = 4
    for site in site_names_in_order:
        ws.cell(row=1, column=col, value=site)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(Q1_YEAR_COLS) - 1)
        col += len(Q1_YEAR_COLS)

    # Ligne 2 : années
    col = 4
    for _ in site_names_in_order:
        for annee, _ in Q1_YEAR_COLS:
            ws.cell(row=2, column=col, value=annee)
            col += 1

    # Mise en forme en-têtes
    total_cols = 3 + len(site_names_in_order) * len(Q1_YEAR_COLS)
    for col_idx in range(1, total_cols + 1):
        for row_idx in (1, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            if row_idx == 1 and col_idx >= 4:
                cell.fill = SITE_HEADER_FILL
            else:
                cell.fill = HEADER_FILL

    # 5. Données (matching par clé)
    for i, (libelle_raw, libelle_clean, indent, key) in enumerate(structure):
        excel_row = i + 3
        cat = classify_line(libelle_clean, indent)
        style = CAT_STYLE[cat]
        fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        font = Font(color=style["font"], bold=style["bold"], size=10)

        code_match = re.match(r"^([A-Z0-9]+)\s*-", libelle_clean)
        code = code_match.group(1) if code_match else ""
        if code and libelle_clean.startswith(code):
            libelle_sans_code = libelle_clean[len(code):].lstrip(" -").strip()
        else:
            libelle_sans_code = libelle_clean
        indent_prefix = "  " * indent

        c = ws.cell(row=excel_row, column=1, value=code if code else "—")
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=excel_row, column=2, value=indent_prefix + libelle_sans_code)
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=excel_row, column=3, value=style["label"])
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        col_v = 4
        for site in site_names_in_order:
            line_def = site_lines.get(site, {}).get(key)
            yr_vals = line_def[3] if line_def else {}
            for annee, _ in Q1_YEAR_COLS:
                val = try_num(yr_vals.get(annee))
                cell = ws.cell(row=excel_row, column=col_v, value=val)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = THIN_BORDER
                cell.number_format = "#,##0;-#,##0;—"
                col_v += 1

    # 5. Largeurs et figement
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    for col_idx in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:C{len(structure) + 2}"

    print(f"  [OK] Q1_2026 : {len(structure)} lignes × {len(site_names_in_order)} sites × {len(Q1_YEAR_COLS)} colonnes")


# ─── Sheet KPI_Techniques (depuis CSV) ────────────────────────────────────────
def write_kpi(wb):
    csv_path = os.path.join(BASE, "data_kpi_techniques.csv")
    if not os.path.exists(csv_path):
        print("  [WARN] data_kpi_techniques.csv introuvable, onglet ignoré")
        return
    rows = read_csv(csv_path)
    if not rows:
        return
    ws = wb.create_sheet(title="KPI_Techniques")
    ws.append(rows[0])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for row in rows[1:]:
        ws.append([try_num(v) for v in row])
    ws.freeze_panes = "C2"
    widths = [22, 8, 12, 14, 16, 12, 12, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    print(f"  [OK] KPI_Techniques : {len(rows) - 1} lignes")


# ─── _README ──────────────────────────────────────────────────────────────────
def write_readme(wb):
    ws = wb.create_sheet(title="_README", index=0)

    ws["A1"] = "BDD_dashboard.xlsx"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Base de données unifiée du dashboard P&L Tri CS"
    ws["A2"].font = Font(italic=True, size=11, color="555555")

    ws["A4"] = "Structure du fichier"
    ws["A4"].font = SECTION_FONT
    ws["A5"] = "Onglet"
    ws["B5"] = "Contenu"
    for cell in [ws["A5"], ws["B5"]]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    sheets_desc = [
        ("PL_Annuel",
         "TOUTES les lignes du P&L source, organisées par site et année, color-codées par catégorie. "
         "Reprend exactement la structure du fichier 'Copie de P&L TRI'."),
        ("Q1_2026",
         "Comptes Q1 2026 (jan→mars) : Réel 2025, Réel 2026, Budget 2026."),
        ("KPI_Techniques",
         "KPIs opérationnels : tonnage, dispo, débit, taux de refus, nb trieurs, etc."),
    ]
    for i, (sheet, desc) in enumerate(sheets_desc, start=6):
        ws.cell(row=i, column=1, value=sheet)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 32

    # Légende couleurs
    ws.cell(row=12, column=1, value="Code couleurs PL_Annuel").font = SECTION_FONT

    legend = [
        ("SYNTH",      "Synthèses : CA, PNE, Marge Brute, EBITDA, EBIT"),
        ("PRINCIPAL",  "Lignes affichées dans la vue 'Détails par site' du dashboard"),
        ("SOUS_POSTE", "Sous-postes : drill-down Personnel / Autres coûts"),
        ("INFO",       "Composantes du P&L pour info"),
        ("ANALYTIQUE", "Détail des comptes analytiques (66xxxxx, 665Xxxxx, etc.)"),
    ]
    for i, (cat, desc) in enumerate(legend, start=13):
        style = CAT_STYLE[cat]
        cell_label = ws.cell(row=i, column=1, value=style["label"])
        cell_label.font = Font(color=style["font"], bold=style["bold"], size=11)
        cell_label.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        cell_label.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=2, value=desc)

    ws.cell(row=19, column=1, value="Mise à jour").font = SECTION_FONT
    ws.cell(row=20, column=1,
            value="1. Modifie les valeurs dans les onglets (PL_Annuel, Q1_2026, KPI_Techniques)")
    ws.cell(row=21, column=1, value="2. Enregistre le fichier.")
    ws.cell(row=22, column=1,
            value="3. Lance le script de rafraîchissement pour régénérer le dashboard.")

    ws.cell(row=24, column=1, value="Notes importantes").font = SECTION_FONT
    ws.cell(row=25, column=1,
            value="• Ne renomme pas les onglets ni les colonnes.")
    ws.cell(row=26, column=1,
            value="• Les valeurs €/t sont calculées automatiquement (pas de saisie).")
    ws.cell(row=27, column=1,
            value="• L'indentation des libellés dans PL_Annuel reflète la hiérarchie du P&L source.")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(SRC_FILE):
        print(f"[ERREUR] Fichier source introuvable : {SRC_FILE}")
        return

    wb = Workbook()
    wb.remove(wb.active)

    print("Génération de BDD_dashboard.xlsx...\n")
    write_readme(wb)
    write_pl_annuel(wb)
    write_q1(wb)
    write_kpi(wb)

    wb.save(OUT_FILE)
    size_kb = os.path.getsize(OUT_FILE) // 1024
    print(f"\n[OK] Fichier généré : BDD_dashboard.xlsx ({size_kb} Ko)")


if __name__ == "__main__":
    main()
