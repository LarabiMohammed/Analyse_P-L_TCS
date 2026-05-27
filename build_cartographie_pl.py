"""
build_cartographie_pl.py
─────────────────────────
Génère 'Cartographie_PL_par_site.xlsx'.

Cartographie comptable par site :
  - Sur quelle ligne tombent les refus (VE000089, VE000093)
  - Sur quelle ligne tombe l'électricité (VE000083 isolée ou noyée)
  - Personnel maintenance (VE000054, VE000055)
  - Cessions internes (VE000045 + sous-codes 66xxxxx)
  - Assurances, Taxes, Pénalités potentielles

Usage : python build_cartographie_pl.py
"""
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "Copie de P&L TRI R2023-R2024-R2025-B2026.xlsx")
OUT = os.path.join(BASE, "Cartographie_PL_par_site.xlsx")

SITES = [
    ("H66NTES281 - 0369-TCS QDR-TRI",     "Le Havre",            "NNO",  "propre"),
    ("H66CAET281 - 9170-CS COUER-TRI",    "Nantes",              "COU",  "multifilière"),
    ("H66HTVP281 - U397-TRI CS AM-TRI",   "Amiens",              "NNO",  "propre"),
    ("H66SVBM281 - U058-BEGLES-TRI",      "Bègles",              "SOU",  "multifilière"),
    ("H66SSMT284 - U446-ECOTRI-TRI",      "Millau",              "SOU",  "propre"),
    ("H66IP15281 - 9809-TRI PARI-TRI",    "Paris 15",            "IDF",  "propre"),
    ("H66CORT282 - U487-TRI IF43-TRI,",   "Saran",               "COU",  "multifilière"),
    ("H66ISVT281 - U701-SEVRAN-TRI",      "Sevran",              "IDF",  "propre"),
    ("H66RZOS282 - 9171-PORTES VA-TRI",   "Portes les Valences", "BARA", "propre"),
    ("H66RSAL281 - U310-ALLIER-TRI",      "Chézy",               "BARA", "propre"),
    ("H66SSMT283 - U446-DEMETER-TRI",     "Montpellier",         "SOU",  "propre"),
]

# Styles
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color="003A63", end_color="003A63", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(bold=True, size=16, color="003A63")
SECTION_FONT = Font(bold=True, size=12, color="003A63")
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
MULTI_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PROPRE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def find_line(rows, code_or_libelle):
    """Trouve la première ligne dont le libellé strippé commence par la clé."""
    for i, row in enumerate(rows, 1):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if s.startswith(code_or_libelle):
            return i, row
    return None, None


def get_subcodes(rows, parent_ve_code):
    """Retourne les sous-codes (avec leur valeur R2025) sous un parent VE code donné."""
    items = []
    in_section = False
    parent_indent = 0
    for i, row in enumerate(rows, 1):
        v = row[0]
        if v is None:
            continue
        s = str(v)
        clean = s.strip()
        n_indent = len(s) - len(s.lstrip(" "))
        if clean.startswith(parent_ve_code):
            in_section = True
            parent_indent = n_indent
            continue
        if in_section:
            if clean and n_indent <= parent_indent:
                in_section = False
                continue
            v25 = num(row[4])
            if v25 is not None and abs(v25) > 100:
                items.append((clean, v25, n_indent))
    return items


def analyze_site(ws):
    """Retourne un dict d'analyse pour un site."""
    rows = list(ws.iter_rows(min_col=1, max_col=6, values_only=True))
    result = {
        "elec_isolee": False, "elec_montant": None,
        "energie_total": None,
        "refus_VE89": None, "refus_VE93": None,
        "pers_maint_int": None, "pers_maint_ext": None,
        "pers_avantages": None,
        "cession_charge_total": None, "cession_produit_total": None,
        "cession_subcodes": [],
        "assurances": None, "taxes": None,
        "autres_couts": None,
        "ca": None, "ebitda": None, "ebit": None, "tonnage": None,
    }

    for i, row in enumerate(rows, 1):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        v25 = num(row[4])

        if s.startswith("[R0100-EW-109] Tonnes entrantes"):
            result["tonnage"] = v25
        elif s.startswith("VE000001 - Chiffre d'affaires"):
            result["ca"] = v25
        elif s.startswith("VE000083"):
            result["elec_isolee"] = True
            result["elec_montant"] = v25
        elif s.startswith("VE000077"):
            result["energie_total"] = v25
        elif s.startswith("VE000089"):
            result["refus_VE89"] = v25
        elif s.startswith("VE000093"):
            result["refus_VE93"] = v25
        elif s.startswith("VE000054"):
            result["pers_maint_int"] = v25
        elif s.startswith("VE000055"):
            result["pers_maint_ext"] = v25
        elif s.startswith("VE000057"):
            result["pers_avantages"] = v25
        elif s.startswith("VE000045"):
            result["cession_charge_total"] = v25
        elif s.startswith("VE000046"):
            result["cession_produit_total"] = v25
        elif s.startswith("VE000103"):
            result["assurances"] = v25
        elif s.startswith("VE000102"):
            result["taxes"] = v25
        elif s.startswith("VE000097"):
            result["autres_couts"] = v25
        elif s == "EBITDA":
            result["ebitda"] = v25
        elif s.startswith("EBIT Courant"):
            result["ebit"] = v25

    result["cession_subcodes"] = get_subcodes(rows, "VE000045")
    return result


def write_synthese(wb, results):
    ws = wb.create_sheet("Synthèse")
    headers = [
        "Site", "Région", "Type",
        "CA (M€)", "EBITDA (M€)", "EBIT (M€)", "Tonnage (kt)",
        "Élec isolée ?", "Élec R2025 (€)",
        "Refus VE000089 (€)", "Refus VE000093 dégrillage (€)",
        "Pers. Maint. interne (€)", "Pers. Maint. externe (€)",
        "Cessions Charge (M€)", "% du CA",
        "Assurances (€)", "Taxes (€)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = THIN
    ws.row_dimensions[1].height = 32

    for site_data in results:
        name, region, typ, st = site_data
        ratio = (abs(st["cession_charge_total"]) / st["ca"] * 100) if (st["cession_charge_total"] and st["ca"]) else None
        ws.append([
            name, region, typ,
            round(st["ca"] / 1e6, 2) if st["ca"] else None,
            round(st["ebitda"] / 1e6, 2) if st["ebitda"] else None,
            round(st["ebit"] / 1e6, 2) if st["ebit"] else None,
            round(st["tonnage"] / 1e3, 1) if st["tonnage"] else None,
            "OUI" if st["elec_isolee"] else "non (noyée)",
            int(st["elec_montant"]) if st["elec_montant"] else None,
            int(st["refus_VE89"]) if st["refus_VE89"] else None,
            int(st["refus_VE93"]) if st["refus_VE93"] else None,
            int(st["pers_maint_int"]) if st["pers_maint_int"] else None,
            int(st["pers_maint_ext"]) if st["pers_maint_ext"] else None,
            round(st["cession_charge_total"] / 1e6, 2) if st["cession_charge_total"] else None,
            round(ratio, 1) if ratio else None,
            int(st["assurances"]) if st["assurances"] else None,
            int(st["taxes"]) if st["taxes"] else None,
        ])
        # Color fill selon type
        row_idx = ws.max_row
        fill = MULTI_FILL if typ == "multifilière" else PROPRE_FILL
        ws.cell(row=row_idx, column=3).fill = fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN
            if col_idx >= 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00;-#,##0.00;—" if col_idx in (4, 5, 6, 14) else "#,##0;-#,##0;—"

    ws.freeze_panes = "D2"
    widths = [22, 8, 14, 12, 12, 12, 11, 13, 14, 16, 18, 18, 18, 15, 10, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_cessions_detail(wb, results):
    ws = wb.create_sheet("Cessions internes - détail")
    ws.append(["Site", "Type", "Code", "Libellé", "Montant R2025 (€)", "% du total"])
    for cell in ws[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = THIN

    for name, region, typ, st in results:
        total_cession = abs(st["cession_charge_total"]) if st["cession_charge_total"] else 0
        if not st["cession_subcodes"]:
            ws.append([name, typ, "—", "Pas de sous-codes significatifs", None, None])
            row_idx = ws.max_row
            fill = MULTI_FILL if typ == "multifilière" else PROPRE_FILL
            ws.cell(row=row_idx, column=2).fill = fill
            continue
        for libelle, montant, indent in st["cession_subcodes"]:
            # Extraire code et libellé
            parts = libelle.split(" - ", 1)
            code = parts[0] if parts else ""
            lib = parts[1] if len(parts) > 1 else ""
            pct = abs(montant) / total_cession * 100 if total_cession > 0 else None
            ws.append([name, typ, code, lib, int(montant), round(pct, 1) if pct else None])
            row_idx = ws.max_row
            fill = MULTI_FILL if typ == "multifilière" else PROPRE_FILL
            ws.cell(row=row_idx, column=2).fill = fill
            ws.cell(row=row_idx, column=5).number_format = "#,##0;-#,##0;—"

    ws.freeze_panes = "A2"
    widths = [22, 14, 14, 60, 18, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_questions(wb, results):
    """Onglet avec les questions ciblées par site (à utiliser dans les mails)."""
    ws = wb.create_sheet("Questions par site", 0)  # En premier
    ws["A1"] = "Questions ciblées par site — à intégrer dans les mails aux responsables"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = ("Ces questions sont générées automatiquement en fonction des particularités identifiées dans la BDD : "
                "lignes manquantes, montants atypiques, postes à clarifier.")
    ws["A3"].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells("A3:E3")

    ws.append([""])
    ws.append(["Site", "Type", "Question(s) à poser"])
    hdr_row = ws.max_row
    for col_idx in range(1, 4):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = THIN

    for name, region, typ, st in results:
        questions = []
        # Q1 : Refus
        if st["refus_VE89"] and abs(st["refus_VE89"]) > 1000:
            questions.append(f"• Le poste 'Traitement et évacuation des sous-produits' affiche {int(st['refus_VE89']):,} € en R2025 : c'est bien le coût de traitement de tes refus ?")
        else:
            questions.append("• Le coût de traitement de tes refus n'apparaît pas isolément sur la ligne VE000089 — sur quelle autre ligne tombe-t-il ? (charges générales ? prestation Veolia interne ?)")

        # Q2 : Électricité
        if st["elec_isolee"]:
            questions.append(f"• Électricité : tu as une ligne VE000083 'Electricité' à {int(st['elec_montant']):,} € en R2025. Penses-tu que ce montant est représentatif d'un coût annuel normal ?")
        else:
            questions.append("• Électricité : pas de ligne 'Electricité' isolée dans ton P&L. Comment ton électricité est-elle comptabilisée ? (refacturation interne ? noyée dans 'Autres énergies' ?)")

        # Q3 : Personnel maintenance
        if st["pers_maint_int"] and abs(st["pers_maint_int"]) > 1000:
            questions.append(f"• Personnel maintenance interne : {int(st['pers_maint_int']):,} € en R2025. C'est cohérent avec ton équipe maintenance dédiée ?")
        else:
            questions.append("• Personnel maintenance interne : ligne VE000054 à zéro chez toi. Tu as bien des techniciens maintenance internes ? Si oui, sur quelle ligne tombe leur coût ?")

        # Q4 : Cessions internes (pour les multi-filières surtout)
        if typ == "multifilière":
            questions.append(f"• Cessions internes : {int(abs(st['cession_charge_total'] or 0)):,} € en R2025 ({(abs(st['cession_charge_total'])/st['ca']*100 if st['ca'] else 0):.1f}% de ton CA). Peux-tu confirmer la composition et nous expliquer comment sont fixés ces prix de cession avec l'UVE jumelle ?")
        elif st["cession_charge_total"] and abs(st["cession_charge_total"]) > 100000:
            questions.append(f"• Cessions internes : {int(abs(st['cession_charge_total'])):,} € en R2025. C'est principalement de la collecte/transport ? Peux-tu confirmer ce que ça recouvre ?")
        else:
            questions.append("• Cessions internes : zéro pour ton site. Confirmes-tu qu'il n'y a aucun flux financier interne Veolia avec d'autres entités ?")

        # Q5 : Autres coûts d'exploitation
        if st["autres_couts"] and abs(st["autres_couts"]) > 500000:
            questions.append(f"• Autres coûts d'exploitation : {int(abs(st['autres_couts'])):,} € en R2025. Peux-tu nous indiquer la nature des principaux montants ? Y a-t-il des éléments exceptionnels (pénalités, perception subventions, C2E, recettes ponctuelles, intéressement non récurrent) ?")

        # Q6 : Pénalités spécifiques
        questions.append("• Pénalités : as-tu eu des pénalités payées ou des avenants en R2025 ? Si oui, montant et nature (taux de captation, dispo, autre) ?")

        # Q7 : C2E et subventions
        questions.append("• Certificats d'économie d'énergie (C2E) ou subventions perçus : y a-t-il eu des recettes liées en R2025 ? Si oui, sur quelle ligne du P&L sont-elles comptabilisées ?")

        # Q8 : Récurrence
        questions.append("• Parmi les éléments ci-dessus, lesquels sont à considérer comme 'exceptionnels' / non répétables en 2026 ?")

        # Assembler les questions en une seule cellule
        q_text = "\n".join(questions)
        ws.append([name, typ, q_text])
        row_idx = ws.max_row
        fill = MULTI_FILL if typ == "multifilière" else PROPRE_FILL
        ws.cell(row=row_idx, column=2).fill = fill
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = 200
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx).border = THIN

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 120
    ws.freeze_panes = "A6"


def write_readme(wb):
    ws = wb.create_sheet("_README", 0)
    ws["A1"] = "Cartographie comptable par site — Centres de Tri CS"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Document de travail pour préparer les échanges avec les responsables de site."
    ws["A3"].font = Font(italic=True, size=11, color="555555")

    ws["A5"] = "Contenu des onglets"
    ws["A5"].font = SECTION_FONT
    onglets = [
        ("Questions par site", "Questions ciblées à poser à chaque responsable (à copier-coller dans les mails)"),
        ("Synthèse", "Tableau récap des KPIs financiers + indicateurs cartographiques (élec isolée, personnel maint, etc.)"),
        ("Cessions internes - détail", "Décomposition ligne par ligne des cessions internes (sous-codes 66xxxxx) avec montants R2025"),
    ]
    for i, (nom, desc) in enumerate(onglets, start=6):
        ws.cell(row=i, column=1, value=nom).font = Font(bold=True, color="003A63")
        ws.cell(row=i, column=2, value=desc)

    ws["A10"] = "Code couleurs"
    ws["A10"].font = SECTION_FONT
    ws.cell(row=11, column=1, value="Multifilière").fill = MULTI_FILL
    ws["B11"] = "Sites avec UVE jumelée : Nantes, Bègles, Saran"
    ws.cell(row=12, column=1, value="En propre").fill = PROPRE_FILL
    ws["B12"] = "Sites standalone : Le Havre, Amiens, Paris 15, Sevran, Chézy, Portes les Valences, Montpellier, Millau"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def main():
    print("Lecture du fichier source...")
    src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    results = []
    for sheet, name, region, typ in SITES:
        if sheet not in src.sheetnames:
            print(f"  [WARN] {name}: onglet introuvable")
            continue
        ws = src[sheet]
        st = analyze_site(ws)
        results.append((name, region, typ, st))
        print(f"  [OK] {name:<22} (cession charge: {st['cession_charge_total']})")
    src.close()

    print("\nGénération du fichier Excel...")
    wb = Workbook()
    wb.remove(wb.active)
    write_readme(wb)
    write_questions(wb, results)
    write_synthese(wb, results)
    write_cessions_detail(wb, results)
    wb.save(OUT)
    size_kb = os.path.getsize(OUT) // 1024
    print(f"\n[OK] Fichier généré : {os.path.basename(OUT)} ({size_kb} Ko)")


if __name__ == "__main__":
    main()
